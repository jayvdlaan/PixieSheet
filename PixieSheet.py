#  Copyright (c) 2022 Teitoku42. All Rights Reserved.
#
#  Permission is hereby granted, free of charge, to any person obtaining a copy
#  of this software and associated documentation files (the "Software"), to deal
#  in the Software without restriction, including without limitation the rights
#  to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
#  copies of the Software, and to permit persons to whom the Software is
#  furnished to do so, subject to the following conditions:
#
#  The above copyright notice and this permission notice shall be included in all
#  copies or substantial portions of the Software.
#
#  THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
#  IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
#  FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
#  AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
#  LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
#  OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
#  SOFTWARE.

import array
import sys

import xlsxwriter
import xlsxwriter.utility
import numpy
import getopt
import openpyxl
import math
from colorsys import rgb_to_hls, hls_to_rgb
from openpyxl.styles.colors import Color, COLOR_INDEX
from PIL import Image, ImageColor

# TODO: Organize functions into classes and cleanup
# TODO: Replace xlsxwriter with openpyxl
# TODO: Move this to a class and seperate python file
RGBMAX = 0xff  # Corresponds to 255
HLSMAX = 240  # MS excel's tint function expects that HLS is base 240. see:


# https://social.msdn.microsoft.com/Forums/en-US/e9d8c136-6d62-4098-9b1b-dac786149f43/excel-color-tint-algorithm-incorrect?forum=os_binaryfile#d3c2ac95-52e0-476b-86f1-e2a697f24969


# Converts rgb values in range (0,1) or a hex string of the form '[#aa]rrggbb' to HLSMAX based HLS (alpha values are
# ignored)
def rgb_to_ms_hls(red, green=None, blue=None):
    if green is None:
        if isinstance(red, str):
            if len(red) > 6:
                red = red[-6:]  # Ignore preceding '#' and alpha values
            blue = int(red[4:], 16) / RGBMAX
            green = int(red[2:4], 16) / RGBMAX
            red = int(red[0:2], 16) / RGBMAX
        else:
            red, green, blue = red

    h, l, s = rgb_to_hls(red, green, blue)
    return int(round(h * HLSMAX)), int(round(l * HLSMAX)), int(round(s * HLSMAX))


# Converts HLSMAX based HLS values to rgb values in the range (0,1)
def ms_hls_to_rgb(hue, lightness=None, saturation=None):
    if lightness is None:
        hue, lightness, saturation = hue

    return hls_to_rgb(hue / HLSMAX, lightness / HLSMAX, saturation / HLSMAX)


# Converts (0,1) based RGB values to a hex string 'rrggbb'
def rgb_to_hex(red, green=None, blue=None):
    if green is None:
        red, green, blue = red
    return ('%02x%02x%02x' % (int(round(red * RGBMAX)), int(round(green * RGBMAX)), int(round(blue * RGBMAX)))).upper()


# Gets theme colors from the workbook
def get_theme_colors(wb):
    # see: https://groups.google.com/forum/#!topic/openpyxl-users/I0k3TfqNLrc
    from openpyxl.xml.functions import QName, fromstring
    xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    root = fromstring(wb.loaded_theme)
    theme_el = root.find(QName(xlmns, 'themeElements').text)
    color_schemes = theme_el.findall(QName(xlmns, 'clrScheme').text)
    first_color_scheme = color_schemes[0]

    colors = []

    for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
        accent = first_color_scheme.find(QName(xlmns, c).text)
        for i in list(accent):  # walk all child nodes, rather than assuming [0]
            if 'window' in i.attrib['val']:
                colors.append(i.attrib['lastClr'])
            else:
                colors.append(i.attrib['val'])

    return colors


# Tints a HLSMAX based luminance
def tint_luminance(tint, lum):
    # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
    if tint < 0:
        return int(round(lum * (1.0 + tint)))
    else:
        return int(round(lum * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint))))


# Given a workbook, a theme number and a tint return a hex based rgb
def theme_and_tint_to_rgb(wb, theme, tint):
    rgb = get_theme_colors(wb)[theme]
    h, l, s = rgb_to_ms_hls(rgb)
    return rgb_to_hex(ms_hls_to_rgb(h, tint_luminance(tint, l), s))


# Loads the passed image into a PIL image object
def load_image(path: str) -> Image:
    image = Image.open(path)
    if image.format != "PNG" or image.mode != "RGB":
        raise RuntimeError("Only non-transparent PNGs are supported")

    return image


# Obtains image starting coordinates and size of pixels
def get_image_info(image: Image) -> dict:
    out_info = {"begin": [], "pixel-size": []}
    header_colors = [[255, 0, 0], [0, 255, 0], [0, 0, 255]]

    # Search image for header pixels (red green and blue)
    header_found = False
    header_check_index = 0
    previous_size = 0
    size_counter = 0
    for y in range(image.height):
        if header_found:
            break

        for x in range(image.width):
            current_pixel = image.getpixel((x, y))
            if header_found:
                break

            # Found initial header pixel
            if numpy.array_equal(current_pixel, header_colors[header_check_index]):
                out_info["begin"] = [x, y]
                for inner_x in range(x, image.width):
                    current_pixel = image.getpixel((inner_x, y))
                    next_pixel = [-1, -1, -1]
                    if inner_x + 1 < image.width:
                        next_pixel = image.getpixel((inner_x + 1, y))

                    # Check if we are still counting the same pixel
                    if numpy.array_equal(current_pixel, header_colors[header_check_index]):
                        size_counter += 1
                        # Check if we're entering a pixel that belongs to the header and is not the final one
                        if header_check_index + 1 < len(header_colors) and \
                                numpy.array_equal(next_pixel, header_colors[header_check_index + 1]):
                            if size_counter != previous_size and header_check_index > 0:
                                raise RuntimeError("Header pixels differ in size")
                            else:
                                previous_size = size_counter
                                size_counter = 0
                                header_check_index += 1
                        # Check if we're entering a pixel that is not part of header colors anymore indicating finish
                        elif header_check_index + 1 == len(header_colors) and not \
                                numpy.array_equal(next_pixel, header_colors[header_check_index]):
                            if size_counter != previous_size:
                                raise RuntimeError("Header pixels differ in size")
                            else:
                                out_info["pixel-size"] = size_counter
                                header_found = True
                                break
                    # Behaviour if pixel is not of the expected header color
                    else:
                        raise RuntimeError("Unexpected pixel color encountered when interpreting header")

    if not header_found:
        raise RuntimeError("Header was not found in image")

    return out_info


# Trims passed image to make starting coordinates 0, 0
def trim_image(image: Image, image_info: dict) -> Image:
    left = image_info["begin"][0]
    top = image_info["begin"][1]
    right = image.width
    bottom = image.height
    return image.crop((left, top, right, bottom))


# Calculate pixel hop count
def calculate_pixel_hops(axis_size: int, pixel_size: int) -> int:
    hops = math.floor(axis_size / pixel_size)
    if axis_size % pixel_size > 0:
        hops += 1

    return int(hops)


# Generate a pixel art matrix from the provided image where a pixel of x size is shrunk to a 1x1 pixel
def generate_pixel_map(image: Image, pixel_size: int) -> numpy.ndarray:
    out_array = []
    width_hops = calculate_pixel_hops(image.width, pixel_size)
    height_hops = calculate_pixel_hops(image.height, pixel_size)
    for y in range(height_hops):
        actual_y = y * pixel_size
        y_row = []
        for x in range(width_hops):
            actual_x = x * pixel_size
            y_row.append(image.getpixel((actual_x, actual_y)))
        out_array.append(y_row)

    return numpy.asarray(out_array, dtype="uint8")


# TODO: Does PIL have a class for this?
def sheet_rgb_to_hex(rgb: tuple):
    return "%02x%02x%02x" % rgb


# TODO: Split functionality into different classes
# Converts a pixel map to a spreadsheet
def map_to_sheet(pixel_map: numpy.ndarray,
                 output_path: str,
                 x_begin: int,
                 y_begin: int,
                 column_width: float,
                 row_height: float):
    book = xlsxwriter.Workbook(output_path)
    sheet = book.add_worksheet("Art")

    width = pixel_map.shape[1]
    height = pixel_map.shape[0]
    corner_offset = 1

    sheet.set_column(0, width, column_width)
    sheet.set_default_row(row_height)

    for x in range(width):
        sheet.write(0, x + corner_offset, x_begin + x)

    for y in range(height):
        sheet.write(y + corner_offset, 0, y_begin + y)

    for y in range(height):
        for x in range(width):
            rgb = pixel_map[y][x]
            fmt = book.add_format({"bg_color": "#" + sheet_rgb_to_hex((rgb[0], rgb[1], rgb[2]))})
            sheet.write(y + corner_offset, x + corner_offset, "", fmt)

    book.close()


# Performs preprocessing on image, then converts it to a spreadsheet
def image_to_sheet(options: dict):
    image = load_image(options["input"])
    image_info = {"begin": (0, 0), "pixel-size": 1}
    if not options["spm"]:
        image_info = get_image_info(image)
        image = trim_image(image, image_info)
    pixel_map = generate_pixel_map(image, image_info["pixel-size"])
    map_to_sheet(pixel_map,
                 options["output"],
                 options["begin"]["x"],
                 options["begin"]["y"],
                 options["dimensions"]["w"],
                 options["dimensions"]["h"])


# Establishes dimension of the image presented in the sheet
def get_sheet_dimensions(sheet) -> dict:
    dimensions = {"x": 0, "y": 0}
    corner_offset = 1
    impl_offset = 1
    index = 0
    while True:
        if sheet.cell(impl_offset, impl_offset + corner_offset + index).value is None:
            break

        index += 1
        dimensions["x"] += 1

    index = 0
    while True:
        if sheet.cell(impl_offset + corner_offset + index, impl_offset).value is None:
            break

        index += 1
        dimensions["y"] += 1

    return dimensions


# Generates a pixel map of the image found in a spreadsheet
def sheet_to_map(workbook, sheet, dimensions: dict) -> numpy.ndarray:
    corner_offset = 1
    impl_offset = 1
    real_offset = corner_offset + impl_offset
    image_map = []
    for y in range(dimensions["y"]):
        row = []
        for x in range(dimensions["x"]):
            cell = sheet.cell(real_offset + y, real_offset + x)
            color = cell.fill.fgColor
            rgb_str = ""
            if color.type == "rgb":
                rgb_str = "#" + color.rgb
            elif color.type == "theme":
                rgb_str = "#FF" + theme_and_tint_to_rgb(workbook, color.theme, color.tint)
            elif color.type == "indexed":
                rgb_str = "#" + COLOR_INDEX[color.index]
            else:
                raise RuntimeError("Unsupported cell color type")

            argb = ImageColor.getcolor(rgb_str, "RGBA")
            row.append((argb[1], argb[2], argb[3]))

        image_map.append(row)

    return numpy.asarray(image_map, dtype="uint8")


# Converts a spreadsheet to a 1x1 pixel image
def sheet_to_image(options: dict):
    book = openpyxl.load_workbook(options["input"])
    sheet = book.active
    dimensions = get_sheet_dimensions(sheet)
    image_map = sheet_to_map(book, sheet, dimensions)
    image = Image.fromarray(image_map)
    image.save(options["output"])


def main(cmdline):
    try:
        options, args = getopt.getopt(cmdline, "hi:o:x:y:l:b:sm:")
    except getopt.GetoptError as err:
        print("PixieSheet.py -h for instructions")
        return

    if len(options) == 0:
        print("PixieSheet.py -h for instructions")
        return

    set_values = {"input": None, "output": None, "mode": None, "begin": {"x": 0, "y": 0},
                  "dimensions": {"w": 2.57, "h": 15.75}, "spm": False}
    for option, argument in options:
        if option == "-i":
            set_values["input"] = argument
        elif option == "-o":
            set_values["output"] = argument
        elif option == "-m":
            set_values["mode"] = argument
        elif option == "-x":
            set_values["begin"]["x"] = int(argument)
        elif option == "-y":
            set_values["begin"]["y"] = int(argument)
        elif option == "-l":
            set_values["dimensions"]["w"] = int(argument)
        elif option == "-b":
            set_values["dimensions"]["h"] = int(argument)
        elif option == "-s":
            set_values["spm"] = True
        else:
            print("PixieSheet - Fast and easy pixel art to spreadsheet\n"
                  "\n"
                  "-m (Required) <Mode> - Specify whether you wanna convert image->spreadsheet or spreadsheet->image "
                  "(values: tosheet, fromsheet)"
                  "-i (Required) <Input path> - Specify the image to turn into a spreadsheet.\n"
                  "-o (Required) <Output name> - Specify the path of the output spreadsheet.\n"
                  "-x <First pixel X> - Specify the X coordinate of the first pixel to place on r/place\n"
                  "-y <First pixel Y> - Specify the Y coordinate of the first pixel to place on r/place\n"
                  "-l <Cell width> - Specify the desired cell width (default is 2.57).\n"
                  "-b <Cell height> - Specify the desired cell height (default is 15.75).\n"
                  "-s <Single pixel mode> - Run the script in single pixel mode, this will assume that your image "
                  "already has 1x1 pixels, this removes the need for a header.\n")
            return

    if set_values["input"] is None or set_values["output"] is None or set_values["mode"] is None:
        print("PixieSheet.py -h for instructions")
        return

    if set_values["mode"] == "tosheet":
        image_to_sheet(set_values)
    elif set_values["mode"] == "fromsheet":
        sheet_to_image(set_values)
    else:
        print("PixieSheet.py -h for instructions")
        return


if __name__ == '__main__':
    main(sys.argv[1:])
